#!/usr/bin/env python
# coding: utf-8

# In[ ]:


__author__ = 'abigail'

"""
Updates age/sex brackets for VMAP 2.0 as participants are enrolled (MAP 338 and higher). Information is pulled from the PTD.
"""

import os
import time
# from send_mail import send_mail
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from redcap import Project

serviceURL = 'https://redcap.vanderbilt.edu/api/'
#map_tracking = Project(serviceURL, os.environ['VMAP_PARTICIPANT_TRACKING'])
#edc = Project(serviceURL, os.environ['VMAP_ELECTRONIC_DATA_CAPTURE'])
#elig_edc = Project(serviceURL, os.environ['VMAP_ELIGIBILITY_EDC'])

# REDCAP REGISTRY REPORT: VMAP 2.0 Eligibility Visits Scheduled
eligibility_scheduled_report = map_tracking.export_report(report_id='360321', format_type='json', raw_or_label='raw',
                                                           raw_or_label_headers='raw',
                                                           export_checkbox_labels=False,
                                                           df_kwargs=None)
exclude = [item for item in eligibility_scheduled_report if
           item['elig_visit_resched'] == '0' and item['elig_visit_no_resched'] != '' or
           item['elig_visit_resched1'] == '0' and item['elig_visit_no_resched1'] != '' or
           item['elig_visit_resched2'] == '0' and item['elig_visit_no_resched2'] != '']
eligibility_scheduled_report = [item for item in eligibility_scheduled_report if item not in exclude]

vmap_edc = edc.export_records(fields=['map_id', 'vf_wrapup_date_time'])
eligibility_edc = elig_edc.export_records(fields=['vmac_id', 'vf_wrapup_date_time'])


def download_ptd():
    """
    downloads enrollment data from the PTD and returns a list of dictionaries for map 338 and higher.
    """
    import pandas as pd
    #get repeating ptd info
    repeat_ptd= map_tracking.export_records(fields=['visit1_date','vmac_id'], events=['enrollmentbaseline_arm_1'])
    #get first instance of repeating scheduling info
    first_repeat_ptd= [i for i in repeat_ptd if i['redcap_repeat_instance']== 1]
    #get static ptd info
    static_ptd = map_tracking.export_records(fields=['map_id', 'dob', 'race', 'sex', 'scd_status', 'vmac_id'], events=['enrollmentbaseline_arm_1'])
    #convert to df and merge
    static_df = pd.DataFrame(static_ptd)
    static_df.drop(['redcap_repeat_instrument', 'redcap_repeat_instance'], axis=1, inplace=True)
    repeating_df = pd.DataFrame(first_repeat_ptd)
    merged_df=static_df.merge(repeating_df, how='outer', on=('vmac_id','redcap_event_name'))
    #convert to list of dictionaries
    all_ptd = merged_df.to_dict('records')
    #select for only vmap 2.0
    ptd_map20 = [item for item in all_ptd if item['map_id'] >= '338']
    return ptd_map20

def age_sex_breakdowns(dataset, visit_type_date):
    """
    calculates age at enrollment using visit1_date variable and dob.
    classifies map ids 338 and higher in respective age and sex groups.
    returns a data_dictionary with the number of participants enrolled grouped by each age and sex.
    :param visit_type_date: either visit 1 or eligibility visit
    :param dataset: list of dictionaries containing enrollment info for map 338 and higher
    """
    breakdown = []
    for record in dataset:
        visit = record[visit_type_date].split("-")
        dob = record['dob'].split("-")
        if len(visit) == 3 and len(dob) == 3:
            year = int(visit[0]) - int(dob[0])
            if int(visit[1]) <= int(dob[1]) and int(visit[2]) < int(dob[2]):
                age = year - 1
            else:
                age = year
            a_dictionary = {'vmac_id': record['vmac_id'], 'sex': record['sex']}
            if 54 >= age >= 50:
                a_dictionary['group'] = '1'
            elif 59 >= age >= 55:
                a_dictionary['group'] = '2'
            elif 64 >= age >= 60:
                a_dictionary['group'] = '3'
            elif 69 >= age >= 65:
                a_dictionary['group'] = '4'
            elif 74 >= age >= 70:
                a_dictionary['group'] = '5'
            elif 79 >= age >= 75:
                a_dictionary['group'] = '6'
            elif 84 >= age >= 80:
                a_dictionary['group'] = '7'
            elif 89 >= age >= 85:
                a_dictionary['group'] = '8'
            elif age >= 90:
                a_dictionary['group'] = '9'
            breakdown.append(a_dictionary)
    male = [item for item in breakdown if item['sex'] == '1']
    female = [item for item in breakdown if item['sex'] == '2']

    my_dictionary = {'group1_male_num': len([item for item in male if item['group'] == '1']),
                     'group1_female_num': len([item for item in female if item['group'] == '1']),
                     'group2_male_num': len([item for item in male if item['group'] == '2']),
                     'group2_female_num': len([item for item in female if item['group'] == '2']),
                     'group3_male_num': len([item for item in male if item['group'] == '3']),
                     'group3_female_num': len([item for item in female if item['group'] == '3']),
                     'group4_male_num': len([item for item in male if item['group'] == '4']),
                     'group4_female_num': len([item for item in female if item['group'] == '4']),
                     'group5_male_num': len([item for item in male if item['group'] == '5']),
                     'group5_female_num': len([item for item in female if item['group'] == '5']),
                     'group6_male_num': len([item for item in male if item['group'] == '6']),
                     'group6_female_num': len([item for item in female if item['group'] == '6']),
                     'group7_male_num': len([item for item in male if item['group'] == '7']),
                     'group7_female_num': len([item for item in female if item['group'] == '7']),
                     'group8_male_num': len([item for item in male if item['group'] == '8']),
                     'group8_female_num': len([item for item in female if item['group'] == '8']),
                     'group9_male_num': len([item for item in male if item['group'] == '9']),
                     'group9_female_num': len([item for item in female if item['group'] == '9'])}
    return my_dictionary


def update_table1(wb, sheet, date, my_dictionary):
    """
    Loads the template excel file, updates the bottom two rows and saves a copy with today's date
    :param my_dictionary: dictionary containing number of participants enrolled grouped by age and sex
    :param date: today's date
    :param wb: active workbook
    :param sheet: active sheet
    """
    vmap_goals = []
    for x in range(3, 21):
        vmap_goals.append((x, sheet.cell(row=6, column=x).value)[1])
    n = 0
    x = 3
    for item in my_dictionary.values():
        sheet.cell(row=6, column=x).value = str(item) + '/' + str(vmap_goals[n])
        legacy_val = sheet.cell(row=3, column=x).value
        if legacy_val is None:
            legacy_val = 0
        sheet.cell(row=7, column=x).value = item + legacy_val
        n = n + 1
        x = x + 1
    total_vmap20 = sum(my_dictionary.values())
    overall_total = total_vmap20 + sheet.cell(row=5, column=21).value
    sheet.cell(row=6, column=21).value = total_vmap20
    sheet.cell(row=7, column=21).value = overall_total
    #wb.save('C:/Users/pechmakr/OneDrive - VUMC/VMAP 2.0 Recruitment/VMAP_Cohort_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx')
    wb.save('/Users/workmeaw/Desktop/VMAP_Enrollment_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx')


def update_race_table(wb, sheet, date, race_dictionary, row_num):
    """
    Loads the template excel file, updates each row and saves a copy with today's date
    :param wb: active workbook
    :param sheet: active sheet
    :param date: today's date
    :param race_dictionary: dictionary filtered by race containing number of participants in each age/sex category
    :param row_num: specifies the row number that we want to fill
    """
    x = 3
    for item in race_dictionary.values():
        sheet.cell(row=row_num, column=x).value = str(item)
        sheet.cell(row=row_num, column=x).alignment = Alignment(horizontal='center')
        x = x + 1
    total_race = sum(race_dictionary.values())
    sheet.cell(row=row_num, column=21).value = total_race
    #wb.save('C:/Users/pechmakr/OneDrive - VUMC/VMAP 2.0 Recruitment/VMAP_Cohort_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx')
    wb.save('/Users/workmeaw/Desktop/VMAP_Enrollment_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx')

def add_totals(wb, sheet, date, my_dictionary, row_num):
    """
    use overall dictionary (i.e., not broken down by race, like table1_dictionary) to fill in last row
    """
    x = 3
    for item in my_dictionary.values():
        if not sheet.cell(row=row_num, column=x).value:
            sheet.cell(row=row_num, column=x).value = str(item)
        else:
            sheet.cell(row=row_num, column=x).value = str(sheet.cell(row=row_num, column=x).value) + '/' + str(item)
        x = x + 1
    #wb.save('C:/Users/pechmakr/OneDrive - VUMC/VMAP 2.0 Recruitment/VMAP_Cohort_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx')
    wb.save('/Users/workmeaw/Desktop/VMAP_Enrollment_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx')

def visits_complete_filter(data_dictionary, database, id, arm):
    """
    this takes a data dictionary and matches the map ids with the ones in vmap edc or matches vmac ids with the ones in eligibility edc database.
    :param data_dictionary: list of dictionaries containing enrollment/eligibility info
    :param database: vmap_edc or eligibility_edc
    :param id: 'map_id' or 'vmac_id'
    :param arm: 'enrollmentbaseline_arm_1' or None
    :return: records in the database specified with visits completed
    """
    data_dictionary_ids = [item[id] for item in data_dictionary]
    if arm is not None:
        matched_ids = [item for item in database if
                       item[id] in data_dictionary_ids and item['redcap_event_name'] == arm]
    else:
        matched_ids = [item for item in database if item[id] in data_dictionary_ids]
    visits_complete = []
    for item in matched_ids:
        if item['vf_wrapup_date_time'] != '':
            visits_complete.append(item[id])
    visits_complete_records = []
    for item in data_dictionary:
        if item[id] in visits_complete:
            visits_complete_records.append(item)
    return visits_complete_records


def run():
    """
    runs all the pieces and sends an email.
    """
    date = time.strftime("%m-%d-%Y")
    ptd_map20 = download_ptd()
    eligibility_data = eligibility_scheduled_report
    # table 1 dictionary
    table1_complete = visits_complete_filter(ptd_map20, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    table1_dictionary = age_sex_breakdowns(table1_complete, 'visit1_date')
    # path = 'C:/Users/pechmakr/Desktop/repos/data_management_reports'
    path = '/Users/workmeaw/Documents/GitHub/data_management_reports'
    os.chdir(path)
    wb = load_workbook('VMAP_Cohort_Sex_Age_Breakdown_Template.xlsx', data_only=True)
    sheet = wb.active
    update_table1(wb, sheet, date, table1_dictionary)

    # table 2 - eligibility data visits scheduled
    table2_dictionary = age_sex_breakdowns(eligibility_data, 'elig_date')
    # filter eligibility data by race
    data_race1_eligibility = [item for item in eligibility_data if item['race'] == '1']
    data_race2_eligibility = [item for item in eligibility_data if item['race'] == '2']
    data_race_other_eligibility = [item for item in eligibility_data if
                                   item['race'] != '1' and item['race'] != '2' and item['race'] != '']
    # row for white caucasian
    dict_race_1_t2 = age_sex_breakdowns(data_race1_eligibility, 'elig_date')
    update_race_table(wb, sheet, date, dict_race_1_t2, 13)
    # row for african american
    dict_race_2_t2 = age_sex_breakdowns(data_race2_eligibility, 'elig_date')
    update_race_table(wb, sheet, date, dict_race_2_t2, 12)
    # row for other race
    dict_race_other_t2 = age_sex_breakdowns(data_race_other_eligibility, 'elig_date')
    update_race_table(wb, sheet, date, dict_race_other_t2, 14)
    # get totals for table 3
    overall_total2 = sheet.cell(row=12, column=21).value + sheet.cell(row=13, column=21).value + sheet.cell(row=14,
                                                                                                            column=21).value
    sheet.cell(row=15, column=21).value = overall_total2
    add_totals(wb, sheet, date, table2_dictionary, 15)

    # table 3 - eligibility visits completed
    data_race1_t3 = visits_complete_filter(data_race1_eligibility, eligibility_edc, 'vmac_id', None)
    data_race2_t3 = visits_complete_filter(data_race2_eligibility, eligibility_edc, 'vmac_id', None)
    data_race3_t3 = visits_complete_filter(data_race_other_eligibility, eligibility_edc, 'vmac_id', None)
    # row for white caucasian
    dict_race1_t3 = age_sex_breakdowns(data_race1_t3, 'elig_date')
    update_race_table(wb, sheet, date, dict_race1_t3, 20)
    # row for african american - data_race2_enrol
    dict_race2_t3 = age_sex_breakdowns(data_race2_t3, 'elig_date')
    update_race_table(wb, sheet, date, dict_race2_t3, 19)
    # row for other race - data_race_other_enrol
    dict_race3_t3 = age_sex_breakdowns(data_race3_t3, 'elig_date')
    update_race_table(wb, sheet, date, dict_race3_t3, 21)
    # get totals for table 3
    overall_total3 = sheet.cell(row=19, column=21).value + sheet.cell(row=20, column=21).value + sheet.cell(row=21,
                                                                                                            column=21).value
    sheet.cell(row=22, column=21).value = overall_total3
    total_complete_3 = visits_complete_filter(eligibility_data, eligibility_edc, 'vmac_id', None)
    table3_dictionary = age_sex_breakdowns(total_complete_3, 'elig_date')
    add_totals(wb, sheet, date, table3_dictionary, 22)

    # table 4 Enrollment visits scheduled
    
    table4_dictionary = age_sex_breakdowns(ptd_map20, 'visit1_date')
    # filter ptd data by race
    data_race1_enrol = [item for item in ptd_map20 if item['race'] == '1']
    data_race2_enrol = [item for item in ptd_map20 if item['race'] == '2']
    data_race_other_enrol = [item for item in ptd_map20 if
                             item['race'] != '1' and item['race'] != '2' and item['race'] != '']
    # row for white caucasian
    dict_race_1_t4 = age_sex_breakdowns(data_race1_enrol, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_1_t4, 27)
    # row for african american
    dict_race_2_t4 = age_sex_breakdowns(data_race2_enrol, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_2_t4, 26)
    # row for other race
    dict_race_other_t4 = age_sex_breakdowns(data_race_other_enrol, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_other_t4, 28)
    # get totals for table 4
    overall_total4 = sheet.cell(row=26, column=21).value + sheet.cell(row=27, column=21).value + sheet.cell(row=28,
                                                                                                            column=21).value
    sheet.cell(row=29, column=21).value = overall_total4
    add_totals(wb, sheet, date, table4_dictionary, 29)

    # table 5 - enrollment visits completed
    data_race1_t5 = visits_complete_filter(data_race1_enrol, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    data_race2_t5 = visits_complete_filter(data_race2_enrol, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    data_race3_t5 = visits_complete_filter(data_race_other_enrol, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    # row for white caucasian - data_race1_enrol
    dict_race1_t5 = age_sex_breakdowns(data_race1_t5, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race1_t5, 34)
    # row for african american - data_race2_enrol
    dict_race2_t5 = age_sex_breakdowns(data_race2_t5, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race2_t5, 33)
    # row for other race - data_race_other_enrol
    dict_race3_t5 = age_sex_breakdowns(data_race3_t5, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race3_t5, 35)
    # get totals for table 5
    overall_total5 = sheet.cell(row=33, column=21).value + sheet.cell(row=34, column=21).value + sheet.cell(row=35,
                                                                                                            column=21).value
    sheet.cell(row=36, column=21).value = overall_total5
    total_complete = visits_complete_filter(ptd_map20, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    table5_dictionary = age_sex_breakdowns(total_complete, 'visit1_date')
    add_totals(wb, sheet, date, table5_dictionary, 36)

    # potential participants row 15 minus row 36, aka table2_dictionary - table5_dictionary
    potential_ptps = {key: table2_dictionary[key] - table5_dictionary.get(key, 0) for key in table2_dictionary}
    add_totals(wb, sheet, date, potential_ptps, 8)
    potential_ptps_total = sheet.cell(row=15, column=21).value - sheet.cell(row=36, column=21).value
    sheet.cell(row=8, column=21).value = potential_ptps_total
    
    # URG VMAP 2.0 (cell V6)
    urg_vmap2 = sheet.cell(row=33, column=21).value + sheet.cell(row=35, column=21).value
    sheet.cell(row=6, column=22).value = urg_vmap2

    # URG Overall Cohort (cell V7)
    sheet.cell(row=7, column=22).value = sheet.cell(row=6, column=22).value + sheet.cell(row=5, column=22).value

    # URG Potential Participants (cell V8): (U12+U14)-(U33+U35))
    urg_scheduled_total = sheet.cell(row=12, column=21).value + sheet.cell(row=14, column=21).value
    urg_completed_total = sheet.cell(row=33, column=21).value + sheet.cell(row=35, column=21).value
    urg_potential_total = urg_scheduled_total - urg_completed_total
    sheet.cell(row=8, column=22).value = urg_potential_total

    # SCD Tables
    scd = [item for item in ptd_map20 if item['scd_status'] == '1']
    # Table 6 - Visit scheduled
    table6_dictionary = age_sex_breakdowns(scd, 'visit1_date')
    # filter ptd data by race
    data_race1_scd = [item for item in scd if item['race'] == '1']
    data_race2_scd = [item for item in scd if item['race'] == '2']
    data_race_other_scd = [item for item in scd if item['race'] != '1' and item['race'] != '2' and item['race'] != '']
    # row for white caucasian
    dict_race_1_t6 = age_sex_breakdowns(data_race1_scd, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_1_t6, 41)
    # row for african american
    dict_race_2_t6 = age_sex_breakdowns(data_race2_scd, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_2_t6, 40)
    # row for other race
    dict_race_other_t6 = age_sex_breakdowns(data_race_other_scd, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_other_t6, 42)
    # get totals for table 6
    overall_total6 = sheet.cell(row=40, column=21).value + sheet.cell(row=41, column=21).value + sheet.cell(row=42,
                                                                                                            column=21).value
    sheet.cell(row=43, column=21).value = overall_total6
    add_totals(wb, sheet, date, table6_dictionary, 43)

    # Table 7 - Visit completed
    data_race1_t7 = visits_complete_filter(data_race1_scd, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    data_race2_t7 = visits_complete_filter(data_race2_scd, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    data_race3_t7 = visits_complete_filter(data_race_other_scd, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    # row for white caucasian - data_race1_scd
    dict_race1_t7 = age_sex_breakdowns(data_race1_t7, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race1_t7, 48)
    # row for african american - data_race2_scd
    dict_race2_t7 = age_sex_breakdowns(data_race2_t7, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race2_t7, 47)
    # row for other race - data_race_other_scd
    dict_race3_t7 = age_sex_breakdowns(data_race3_t7, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race3_t7, 49)
    # get totals for table 7
    overall_total7 = sheet.cell(row=47, column=21).value + sheet.cell(row=48, column=21).value + sheet.cell(row=49,
                                                                                                            column=21).value
    sheet.cell(row=50, column=21).value = overall_total7
    total_complete = visits_complete_filter(scd, vmap_edc, 'map_id', 'enrollmentbaseline_arm_1')
    table7_dictionary = age_sex_breakdowns(total_complete, 'visit1_date')
    add_totals(wb, sheet, date, table7_dictionary, 50)
    
    # Table 8 - Brain Donation Race
    # filter brain donation yes data by race
    brain_donation_yes = [item for item in eligibility_data if item['consent_tissue_donation'] == '1']
    brain_donation_no = [item for item in eligibility_data if item['consent_tissue_donation'] == '0']
    brain_donation_dictionary_yes = age_sex_breakdowns(brain_donation_yes, 'visit1_date')
    brain_donation_dictionary_no = age_sex_breakdowns(brain_donation_no, 'visit1_date')
    
    data_race1_donation_yes = [item for item in brain_donation_yes if item['race'] == '1']
    data_race2_donation_yes = [item for item in brain_donation_yes if item['race'] == '2']
    data_race3_donation_yes = [item for item in brain_donation_yes if item['race'] != '1' 
                               and item['race'] != '2' and item['race'] != '']

    data_race1_donation_no = [item for item in brain_donation_no if item['race'] == '1']
    data_race2_donation_no = [item for item in brain_donation_no if item['race'] == '2']
    data_race3_donation_no = [item for item in brain_donation_no if item['race'] != '1' 
                               and item['race'] != '2' and item['race'] != '']

    # row for white caucasian
    dict_race_1_t8_yes = age_sex_breakdowns(data_race1_donation_yes, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_1_t8_yes, 56)
    dict_race_1_t8_no = age_sex_breakdowns(data_race1_donation_no, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_1_t8_no, 56)
    # row for african american
    dict_race_2_t8_yes = age_sex_breakdowns(data_race2_donation_yes, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_2_t8_yes, 55)
    dict_race_2_t8_no = age_sex_breakdowns(data_race2_donation_no, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_2_t8_no, 55)
    # row for other race
    dict_race_3_t8_yes = age_sex_breakdowns(data_race3_donation_yes, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_3_t8_yes, 57)
    dict_race_3_t8_no = age_sex_breakdowns(data_race3_donation_no, 'visit1_date')
    update_race_table(wb, sheet, date, dict_race_3_t8_no, 57)

    # get totals for table 8
    overall_total8_y = 0
    overall_total8_n = 0
    for i in range(55, 58):
        y_value = sheet.cell(row=i, column=21).value.split('/')[0]
        if not y_value:
            y_value = 0
        n_value = sheet.cell(row=i, column=21).value.split('/')[1]
        if not n_value:
            n_value = 0
        overall_total8_y += int(y_value)
        overall_total8_n += int(n_value)
    overall_total8 = '{}/{}'.format(overall_total8_y, overall_total8_n)
    sheet.cell(row=58, column=21).value = overall_total8
    add_totals(wb, sheet, date, brain_donation_dictionary_yes, 58)
    add_totals(wb, sheet, date, brain_donation_dictionary_no, 58)
    
    #subject = 'Weekly VMAP Enrollment Breakdown Report'
    #body = 'All, please find attached the weekly VMAP enrollment breakdown report.'
    #print(body)
    #file = ['C:/Users/pechmakr/Desktop/Reports/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx']

    #file = ['C:/Users/pechmakr/Box Sync/VMAC - Vanderbilt Memory & Aging Project/VMAP - Parent Study/VMAP 2.0 Recruitment/VMAP_Cohort_Breakdown/VMAP_Cohort_Sex_Age_Breakdown_' + date + '.xlsx']
    #to_emails = ['kimberly.r.pechman@vumc.org','angela.jefferson@vumc.org','katie.gifford@vumc.org','michelle.houston@vumc.org',
    #             'jenna.boue@vumc.org','raymond.romano@vumc.org','paige.e.crepezzi@vumc.org','daniel.c.ibarra-scurr@vumc.org']
    # to_emails = ['kimberly.r.pechman@vumc.org']
    #send_mail(to_emails, subject, body, file)


if __name__ == '__main__':
    run()

